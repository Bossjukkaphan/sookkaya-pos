"""อ่านต้นทุนวัสดุต่อเมนูจากชีท 'ต้นทุน' แล้วสร้าง SQL

ชื่อเมนูใน Excel เว้นวรรคต่างจากในฐานข้อมูล จึงต้อง normalize ก่อนจับคู่
เมนูที่ระบุ 2 ค่า (เช่น '135.9/178.2') ใช้ค่าเฉลี่ยตามที่ตกลงใน spec
"""
import openpyxl, re, unicodedata, pathlib

XLSX = "/Users/jw/Downloads/Final_SOOKKAYA_บันทึกรับจ่าย_v15_Latest 3_5_69.xlsx"
OUT = pathlib.Path(__file__).parent / "material-cost.sql"


def norm(s):
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(s)).strip()).lower()


# ชื่อในไฟล์เก่า -> ชื่อในฐานข้อมูล (เฉพาะที่สะกดต่างกัน)
ALIAS = {
    norm("ทรีตเมนต์ขัดผิว และนวดน้ำมันหอมระเหย 90 นาที"): "ทรีตเมนต์ขัดผิว + นวดน้ำมัน 90 นาที",
    norm("ทรีตเมนต์ขัดผิว และนวดน้ำมันหอมระเหย 120 นาที"): "ทรีตเมนต์ขัดผิว + นวดน้ำมัน 120 นาที",
}

# ชื่อในฐานข้อมูลจริง ใช้จับคู่แบบ normalize
DB_NAMES = [
    "นวดแผนไทย 60 นาที", "นวดแผนไทย 90 นาที", "นวดแผนไทย 120 นาที",
    "นวดไทยด้วยบาล์มหรือน้ำมัน 60 นาที", "นวดไทยด้วยบาล์มหรือน้ำมัน 90 นาที",
    "นวดไทยด้วยบาล์มหรือน้ำมัน 120 นาที",
    "นวดคอ บ่า ไหล่ 60 นาที",
    "นวดฝ่าเท้า 60 นาที", "นวดฝ่าเท้า 90 นาที", "นวดฝ่าเท้า 120 นาที",
    "นวดน้ำมันอโรมา 60 นาที", "นวดน้ำมันอโรมา 90 นาที", "นวดน้ำมันอโรมา 120 นาที",
    "นวดคลายกล้ามเนื้อลึก 60 นาที", "นวดคลายกล้ามเนื้อลึก 90 นาที",
    "นวดคลายกล้ามเนื้อลึก 120 นาที",
    "นวดแก้ออฟฟิศซินโดรม 60 นาที", "นวดแก้ออฟฟิศซินโดรม 90 นาที",
    "นวดแก้ออฟฟิศซินโดรม 120 นาที",
    "ทรีตเมนต์ขัดผิว 60 นาที", "ทรีตเมนต์ขัดผิว + นวดน้ำมัน 90 นาที",
    "ทรีตเมนต์ขัดผิว + นวดน้ำมัน 120 นาที",
    "นวดคลายเท้า & คอบ่าไหล่ 60 นาที", "นวดคลายเท้า & คอบ่าไหล่ 90 นาที",
    "นวดคลายเท้า & คอบ่าไหล่ 120 นาที",
    "นวดน้ำมันอุ่น 60 นาที", "นวดน้ำมันอุ่น 90 นาที", "นวดน้ำมันอุ่น 120 นาที",
    "นวดศีรษะดั้งเดิม 60 นาที", "นวดศีรษะดั้งเดิม 90 นาที", "นวดศีรษะดั้งเดิม 120 นาที",
    "นวดศีรษะน้ำมันมะพร้าว 60 นาที", "นวดศีรษะน้ำมันมะพร้าว 90 นาที",
    "นวดศีรษะน้ำมันมะพร้าว 120 นาที",
]
BY_NORM = {norm(n): n for n in DB_NAMES}


def to_cost(v):
    """คืนค่าเฉลี่ยถ้าเป็นรูปแบบ 'a/b' มิฉะนั้นคืนตัวเลขตรงๆ"""
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, str) and "/" in v:
        parts = [float(p) for p in v.split("/") if p.strip()]
        return round(sum(parts) / len(parts), 2)
    return None


wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["ต้นทุน"]

pairs, skipped, unmatched = [], [], []
for r in range(2, ws.max_row + 1):
    raw_name = ws.cell(r, 1).value
    if not raw_name:
        continue
    cost = to_cost(ws.cell(r, 4).value)
    if cost is None:
        skipped.append(str(raw_name).strip())
        continue

    n = norm(raw_name)
    db_name = ALIAS.get(n) or BY_NORM.get(n)
    if db_name is None:
        unmatched.append(str(raw_name).strip())
        continue
    pairs.append((db_name, cost))

values = ",\n".join(f"({cost}, '{name}')" for name, cost in pairs)
OUT.write_text(
    f"""update public.services s
set material_cost = v.cost
from (values
{values}
) as v(cost, name)
where s.name = v.name;
"""
)

print(f"จับคู่ได้ {len(pairs)} เมนู · ไม่มีต้นทุน {len(skipped)} · จับคู่ชื่อไม่ได้ {len(unmatched)}")
if unmatched:
    print("  จับคู่ไม่ได้:", unmatched)
print(f"เขียนไปที่ {OUT}")

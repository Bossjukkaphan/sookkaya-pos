"""ดึงเลขที่ใบเสร็จกับค่าเวลาดิบจากชีทบันทึกขาย ออกเป็น JSON
ไม่แปลงค่าใดๆ ทั้งสิ้น — การตีความเป็นหน้าที่ของ build-sale-time-sql.ts"""
import openpyxl, json, pathlib, datetime
from collections import Counter

XLSX = "/Users/jw/Downloads/Final_SOOKKAYA_บันทึกรับจ่าย_v15_Latest 3_5_69.xlsx"
OUT = pathlib.Path(__file__).parent / "sale-times-raw.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["บันทึกขาย"]

seen, rows, skipped = Counter(), [], 0
for r in range(3, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 8)]
    if row[0] is None and row[3] is None and row[6] is None:
        continue
    receipt = str(row[2]).strip() if row[2] else None
    if not receipt:
        skipped += 1
        continue
    seen[receipt] += 1
    if seen[receipt] > 1:                     # ตอน import เลขซ้ำถูกต่อท้าย -2
        receipt = f"{receipt}-{seen[receipt]}"

    raw = row[1]
    if isinstance(raw, datetime.time):
        raw = raw.strftime("%H:%M")
    elif isinstance(raw, datetime.datetime):
        raw = raw.strftime("%H:%M")
    rows.append({"receipt_no": receipt, "raw": raw})

OUT.write_text(json.dumps(rows, ensure_ascii=False))
print(f"อ่านได้ {len(rows)} แถว · ไม่มีเลขใบเสร็จ {skipped} แถว")
print(f"เขียนไปที่ {OUT}")

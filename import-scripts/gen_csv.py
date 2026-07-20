"""สร้าง CSV สำหรับ upload เข้า staging tables ผ่าน PostgREST"""
import openpyxl, re, datetime, unicodedata, pathlib, csv, json
from collections import Counter, defaultdict

P = "/Users/jw/Downloads/Final_SOOKKAYA_บันทึกรับจ่าย_v15_Latest 3_5_69.xlsx"
OUT = pathlib.Path(__file__).parent / "csv"
OUT.mkdir(exist_ok=True)
wb = openpyxl.load_workbook(P, data_only=True)

def norm(s):
    if s is None: return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(s)).strip()).lower()

def txt(v):
    return "" if v is None else str(v).strip()

def numv(v, default=""):
    if v is None or v == "": return default
    try:
        f = float(v); return str(int(f)) if f == int(f) else str(f)
    except (TypeError, ValueError): return default

def d(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime("%Y-%m-%d")
    return ""

def parse_time(v):
    if isinstance(v, datetime.time): return v.strftime("%H:%M")
    if isinstance(v, datetime.datetime): return v.strftime("%H:%M")
    if isinstance(v, float) and 0 < v < 1:
        m = round(v * 1440); return f"{m//60:02d}:{m%60:02d}"
    return ""

SEEDED = ['นวดแผนไทย 60 นาที','นวดแผนไทย 90 นาที','นวดแผนไทย 120 นาที',
 'นวดไทยด้วยบาล์มหรือน้ำมัน 60 นาที','นวดไทยด้วยบาล์มหรือน้ำมัน 90 นาที','นวดไทยด้วยบาล์มหรือน้ำมัน 120 นาที',
 'นวดคอ บ่า ไหล่ 60 นาที','นวดฝ่าเท้า 60 นาที','นวดฝ่าเท้า 90 นาที','นวดฝ่าเท้า 120 นาที',
 'นวดน้ำมันอโรมา 60 นาที','นวดน้ำมันอโรมา 90 นาที','นวดน้ำมันอโรมา 120 นาที',
 'นวดคลายกล้ามเนื้อลึก 60 นาที','นวดคลายกล้ามเนื้อลึก 90 นาที','นวดคลายกล้ามเนื้อลึก 120 นาที',
 'นวดแก้ออฟฟิศซินโดรม 60 นาที','นวดแก้ออฟฟิศซินโดรม 90 นาที','นวดแก้ออฟฟิศซินโดรม 120 นาที',
 'ทรีตเมนต์ขัดผิว 60 นาที','ทรีตเมนต์ขัดผิว + นวดน้ำมัน 90 นาที','ทรีตเมนต์ขัดผิว + นวดน้ำมัน 120 นาที',
 'นวดคลายเท้า & คอบ่าไหล่ 60 นาที','นวดคลายเท้า & คอบ่าไหล่ 90 นาที','นวดคลายเท้า & คอบ่าไหล่ 120 นาที',
 'นวดน้ำมันอุ่น 60 นาที','นวดน้ำมันอุ่น 90 นาที','นวดน้ำมันอุ่น 120 นาที',
 'นวดศีรษะดั้งเดิม 60 นาที','นวดศีรษะดั้งเดิม 90 นาที','นวดศีรษะดั้งเดิม 120 นาที',
 'นวดศีรษะน้ำมันมะพร้าว 60 นาที','นวดศีรษะน้ำมันมะพร้าว 90 นาที','นวดศีรษะน้ำมันมะพร้าว 120 นาที']
seed_by_norm = {norm(s): s for s in SEEDED}
ALIAS = {norm('ทรีตเมนต์ขัดผิว และนวดน้ำมันหอมระเหย 90 นาที'): 'ทรีตเมนต์ขัดผิว + นวดน้ำมัน 90 นาที',
         norm('ทรีตเมนต์ขัดผิว และนวดน้ำมันหอมระเหย 120 นาที'): 'ทรีตเมนต์ขัดผิว + นวดน้ำมัน 120 นาที'}
VALID_PAY = {'QR Code','บัตรเครดิต','Gowabi','KOL','Member Credit','เงินสด'}
report = {}

def write(name, header, rows):
    with (OUT / name).open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(header); w.writerows(rows)
    return len(rows)

# ---------- customers ----------
ws = wb['ข้อมูลลูกค้า (CRM)']
custs = [(r, [ws.cell(r, c).value for c in range(1, 16)])
         for r in range(4, ws.max_row + 1) if ws.cell(r, 2).value]
by_name = defaultdict(list)
for r, v in custs: by_name[norm(v[1])].append((r, v))

rows = [[txt(v[1]), txt(v[2]), txt(v[3]), txt(v[4]), d(v[5]),
         'สมาชิก' if v[6] and '💳' in str(v[6]) else 'ลูกค้าทั่วไป',
         txt(v[13]), f"xls:{r}"] for r, v in custs]
report['customers'] = write("customers.csv",
    ["name","nickname","phone","line_id","birthday","customer_type","notes","legacy_ref"], rows)

# ---------- sales ----------
ws = wb['บันทึกขาย']
sales = []
for r in range(3, ws.max_row + 1):
    v = [ws.cell(r, c).value for c in range(1, 20)]
    if v[0] is None and v[3] is None and v[6] is None: continue
    sales.append(v)

seen, rows, skipped = Counter(), [], 0
for s in sales:
    date = d(s[0])
    if not date: skipped += 1; continue
    rc = txt(s[2]) or ""
    if rc:
        seen[rc] += 1
        if seen[rc] > 1: rc = f"{rc}-{seen[rc]}"

    ref, nm = "", norm(s[3])
    if nm:
        cands = by_name.get(nm, [])
        if len(cands) == 1: ref = f"xls:{cands[0][0]}"
        elif len(cands) > 1:
            ph = txt(s[4])
            hit = [c for c in cands if txt(c[1][3]) == ph and ph]
            if len(hit) == 1: ref = f"xls:{hit[0][0]}"

    raw = txt(s[6])
    match = seed_by_norm.get(norm(raw)) or ALIAS.get(norm(raw)) or ""
    pay = txt(s[12]) if txt(s[12]) in VALID_PAY else 'ไม่ระบุ'
    net = numv(s[10], "0")
    ms = txt(s[15]) if txt(s[15]) and txt(s[15]) != '—' else ""

    rows.append([rc, date, parse_time(s[1]), ref, txt(s[3]), txt(s[4]), txt(s[5]),
        match, raw, numv(s[7], "0"), txt(s[8]), numv(s[9], "0"), net, numv(s[11], "0"),
        pay, "true" if s[13] else "false", numv(s[14], "0"), ms,
        net if pay == 'Member Credit' else "0",
        numv(s[18], "0") if pay == 'Member Credit' else "0",
        numv(s[17], net) if pay == 'Member Credit' else net])

report['sales'] = write("sales.csv",
    ["receipt_no","sale_date","sale_time","cust_ref","customer_name","customer_phone",
     "th_name","svc_match","svc_raw","price_normal","coupon_promo","discount","net_amount",
     "commission","payment_method","is_request","request_fee","member_status",
     "credit_used","bonus_used","revenue_recognize"], rows)
report['sales_skipped_no_date'] = skipped

# ---------- expenses ----------
ws = wb['รายจ่าย']
rows = []
for r in range(3, ws.max_row + 1):
    v = [ws.cell(r, c).value for c in range(1, 7)]
    if not (v[1] or v[3]) or not d(v[0]): continue
    rows.append([d(v[0]), txt(v[1]) or 'ไม่ระบุ', txt(v[2]) or 'อื่นๆ',
                 numv(v[3], "0"), txt(v[4]), txt(v[5])])
report['expenses'] = write("expenses.csv",
    ["expense_date","item","category","amount","paid_by","notes"], rows)

# ---------- topups ----------
ws = wb['💳 Member Topup']
rows, unmatched = [], []
for r in range(4, ws.max_row + 1):
    v = [ws.cell(r, c).value for c in range(1, 12)]
    if not v[2]: continue
    cands = by_name.get(norm(v[2]), [])
    ref = ""
    if len(cands) == 1: ref = f"xls:{cands[0][0]}"
    elif len(cands) > 1:
        ph = txt(v[3]); hit = [c for c in cands if txt(c[1][3]) == ph and ph]
        if len(hit) == 1: ref = f"xls:{hit[0][0]}"
    if not ref: unmatched.append(txt(v[2])); continue
    t = str(v[4]); tier = 'Silver' if 'Silver' in t else 'Gold' if 'Gold' in t else 'Platinum'
    rows.append([d(v[1]), ref, tier, txt(v[5]) or 'QR Code',
                 numv(v[6], "0"), numv(v[7], "0"), numv(v[8], "0"), d(v[10]), txt(v[9])])
report['topups'] = write("topups.csv",
    ["topup_date","cust_ref","tier","payment_method","cash_received",
     "credit_added","bonus_added","expiry_date","notes"], rows)
report['topups_unmatched'] = unmatched

print(json.dumps(report, ensure_ascii=False, indent=2))
for f in sorted(OUT.iterdir()):
    print(f"  {f.name}  {f.stat().st_size:,} bytes")
